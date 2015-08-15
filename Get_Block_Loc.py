from openpyxl import Workbook
from openpyxl import load_workbook
from geopy.geocoders import Nominatim
import csv
import os

Geocoder = Nominatim()
wb = Workbook()
ws=wb.active
alph_numeric_list=["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]

source  ='result.xlsx'
blocks_loc_info = []
Streets = {}
help_file = 'Block_lat_long.csv'

fr = open(help_file)
read = csv.reader(fr)

for row in read:
    Streets[str(row[0][:row[0].rfind(" ")])]=[row[1],row[2]]
print "Preparing Block and location information file"

dup_check = {}

workbook = load_workbook(source)
work_sheet = workbook.active
max_rows = work_sheet.get_highest_row()

for index in range(0,max_rows+1):
    index+=1
    try:
        block= work_sheet['A'+str(index)].internal_value
        block_name = [block[block.find(" ")+1:]][0][:[block[block.find(" ")+1:]][0].rfind(" ")]
        if(dup_check.has_key(block_name)):
            continue
        row = [block_name, Streets[block_name][0] , Streets[block_name][1] ]
        blocks_loc_info.append(row)
        dup_check[block_name] =0
    except Exception:
        continue
print "Writing to File"

index = 1
for row in blocks_loc_info:
    ws["A"+str(index)]=row[0]
    ws["B"+str(index)]=row[1]
    ws["C"+str(index)]=row[2]
    index+= 1
wb.save('Block_loc.xlsx')

print "Completed"

fr.close()
