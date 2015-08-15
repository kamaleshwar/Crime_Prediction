#authors: Dinesh Gudi & Kamaleshwar

from __future__ import division
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import glob,os
from os.path import basename
import shutil

files=[]

files=glob.glob("Final_data_set\*.xlsx")
accuracy=0
precision=0
recall=0

def calculateIndex(percent_split,ws):
	max_rows = ws.get_highest_row()
	indx = int((float(percent_split)/100)*max_rows)
	return indx

def predictCompare(indx,ws):
	alph_numeric_list=["C","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
	max_rows = ws.get_highest_row()
	max_columns = ws.get_highest_column()
	FP=0
	TP=0
	TN=0
	FN=0
	for index in range(indx,max_rows+1):
		coordinate1=alph_numeric_list[max_columns-4]+str(index)
		coordinate2=alph_numeric_list[max_columns-5]+str(index)
#		print ws[coordinate1].internal_value
#		print ws[coordinate2].internal_value
		x=ws[coordinate1].internal_value
		y=ws[coordinate2].internal_value
		if(x==y and x=="critical"):
			TP=TP+1
		if(x==y and x=="non-critical"):
			TN=TN+1
		if(x!=y and y=="critical"):
			FP=FP+1
		if(x!=y and y=="non-critical"):
			FN=FN+1
	return(TP,FP,TN,FN,max_rows-indx+1)

def naiveBayes(percent_split,file):
	global accuracy
	global precision
	global recall
	wb2 = load_workbook(file)
	ws=wb2.active
	wb3 = load_workbook(file)
	ws1=wb3.active
	max_rows = ws.get_highest_row()
	max_columns = ws.get_highest_column()
	indx = calculateIndex(percent_split,ws)+1
	if(indx>=max_rows):
		indx=indx-1
	for index in range(indx,max_rows+1):
		predict(index,ws,ws1,file)
	wb3.save('results.xlsx')
	(TP,FP,TN,FN,sum) = predictCompare(indx,ws1)
#	print (TP,FP,TN,FN,sum)
#	print max_rows
	accuracy=accuracy+float(TP+TN)/sum
	try:
		recall=recall+float(TP)/(TP+FN)
	except:
		recall=recall+1
	try:
		precision=precision+float(TP)/(FP+TP)
	except:
		precision=precision+1
	return (accuracy,precision,recall)

def predict(row,ws,ws1,file):
	max_rows = ws.get_highest_row()
	max_columns = ws.get_highest_column()
	alph_numeric_list=["C","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
	indx=alph_numeric_list[max_columns-3]+str(row)
#	print indx
	key=ws["A"+str(row)].internal_value+":"+ws["B"+str(row)].internal_value
#	print key
	q_attr=get_q_attr(file,key)
#	print q_attr[0].internal_value
	prob=findProb(q_attr,file)
	ws1[indx]=prob
#	print ws1[indx].internal_value
	return

def get_q_attr(file,key1):
	workbook=load_workbook(file)
	worksheet=workbook.active
	alph_numeric_list=["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
	dict_file={}
	max_rows = worksheet.get_highest_row()
	for index in range(1,max_rows+1):
		month_column="A"
		year_column="B"
		month_row=index
		year_row=index
		month_coordinate=month_column+str(month_row)
		year_coordinate=year_column+str(year_row)
		key=worksheet[month_coordinate].internal_value+":"+str(worksheet[year_coordinate].internal_value)
		value_crude=tuple(worksheet.iter_rows("A"+str(index)+":"+"I"+str(index)))
		crude=value_crude[0]
		value_final=[crude[2],crude[4],crude[5],crude[6],crude[7],crude[8]]
		dict_file[key]=value_final
	return dict_file[key1]

#used for concatenating two columns
def concat(x,y):
	return (x[0].internal_value,y[0].internal_value)

#counts the truth
def count_num(w_play,outcome,file):
	wb2 = load_workbook(file)
	ws=wb2.active
	max_column = ws.get_highest_column()
	count=0
	for play in w_play:
		if play[0].internal_value==outcome:
			count=count+1
	return count

#calculates the probability of the of positive and negetive cases
def findProb(row,file):
    wb2 = load_workbook(file)
    ws=wb2.active
    max_rows = ws.get_highest_row()
    max_columns = ws.get_highest_column()
    column_count = 1
    w_list = []
    alph_numeric_list=["C","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
    word_count_list = []
    global attribute_prob
    attribute_prob=[]
    for index in range(0,max_columns-3):
        word_count_list.append(index)
        attribute_prob.append(index)
    for index in range(1,max_columns-2):
        w_list.append(tuple(ws.iter_rows(alph_numeric_list[index-1]+str(2)+":"+alph_numeric_list[index-1]+str(max_rows))))

    num_yes=count_num(w_list[max_columns-4],"critical",file)
    num_no=count_num(w_list[max_columns-4],"non-critical",file)

    yes_prob=num_yes/float(num_yes+num_no)
    no_prob=num_no/float(num_yes+num_no)

    for index in range(1,max_columns-2):
        w_list[index-1]=map(concat,w_list[index-1],w_list[max_columns-4]);
        word_count_list[index-1]={}

    for index in range(1,max_columns-2):
        for word in w_list[index-1]:
            if not word in word_count_list[index-1].keys():
                word_count_list[index-1][word] = 1
            else:
                word_count_list[index-1][word] = word_count_list[index-1][word] + 1

    for index in range(1,max_columns-2):
        q_attr=row[index-1].coordinate       #"A"+str(row)
#        print q_attr
        try:
		    prob_each_attr_no=word_count_list[index-1][(ws[q_attr].internal_value,"non-critical")]/float(num_no)
        except KeyError:
			prob_each_attr_no=0
        try:
			prob_each_attr_yes=word_count_list[index-1][(ws[q_attr].internal_value,"critical")]/float(num_yes)
        except KeyError:
            prob_each_attr_yes=0
        small_list=[]
        small_list.append(prob_each_attr_no)
        small_list.append(prob_each_attr_yes)
        attribute_prob[index-1]=small_list
#        column_count = column_count+1
#        print small_list
#        print prob_each_attr_no
#        print prob_each_attr_yes
    prob_yes=1
    prob_no=1

    for index in range(1,max_columns-3):
        prob_yes=attribute_prob[index-1][1]*prob_yes
        prob_no=attribute_prob[index-1][0]*prob_no
    prob_yes=prob_yes*yes_prob
    prob_no=prob_no*no_prob
#    print str(prob_yes)+","+str(prob_no)
    return"critical" if prob_yes>prob_no else "non-critical"

#main method
def getPrediction():
	index=1
	wb = Workbook()
	ws=wb.active
	key="April:2015"
	for file in files:
		q_attr=get_q_attr(file,key)
		prediction=findProb(q_attr,file)
		if prediction=="critical":
			result=[basename(os.path.splitext(file)[0]),prediction]
			ws["A"+str(index)]=result[0]
			ws["B"+str(index)]=result[1]
			index=index+1
	try:
		os.remove("result.xlsx")
	except OSError:
		pass
	wb.save("result.xlsx")

def findPrediction():
    global precision
    global accuracy
    global recall
    index=1
    wb = Workbook()
    ws=wb.active
    for file in files:
        (accuracy,precision,recall)=naiveBayes(66,file)
    print "precision="+str(precision/len(files))
    print "accuracy="+str(accuracy/len(files))
    print "recall="+str(recall/len(files))

print "Predicting the Critical areas..."

getPrediction()
print "Prediction completed"

#print "Calculating accuracy precision details..."
findPrediction()

