import pygmaps
import os
import urllib2
import webbrowser
from openpyxl import load_workbook
from openpyxl import Workbook
from geopy.geocoders import Nominatim

geolocator = Nominatim()
count =0
source = 'Block_loc.xlsx'
mymap = pygmaps.maps(41.8369,-87.6847, 12)
wb = load_workbook(source)
ws = wb.active
max_rows = ws.get_highest_row()

def fill_list():
    result = []
    for index in range(1,max_rows + 1):
        result.append([ws['B'+ str(index)].internal_value , ws['C'+str(index)].internal_value] )
    return result

def add_point(lat,lon):
    global mymap
    mymap.addpoint(lat, lon)
    mymap.addradpoint(lat,lon, 500, "#FF0000")

def get_blocks():
    global result
    global count
    result = fill_list()
    for block in result:
        add_point(float(block[0]),float(block[1]))


def draw():
    global mymap
    mymap.draw('./Output.html')

def main():
  print "Gathering all Critical Blocks..."
  get_blocks()
  print "Forming Map..."
  draw()
  print "Map file prepared"
  dir_path = os.getcwd()
  url = 'file:///'+dir_path+'/Output.html'
  webbrowser.open(url)

if __name__ == '__main__':
  main()



